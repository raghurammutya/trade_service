import pandas as pd
import logging
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from enum import Enum

logger = logging.getLogger(__name__)

class SegmentType(Enum):
    EQUITY = "EQ"
    FUTURES_OPTIONS = "FO"
    COMMODITIES = "COM"

class TradebookParser:
    """Enhanced parser for Zerodha tradebook Excel files with auto-detection for EQ/FO/COM"""
    
    # Expected columns in all tradebooks
    BASE_COLUMNS = [
        'Symbol', 'ISIN', 'Trade Date', 'Exchange', 'Segment', 'Series',
        'Trade Type', 'Auction', 'Quantity', 'Price', 'Trade ID', 
        'Order ID', 'Order Execution Time'
    ]
    
    # Additional columns for derivatives (FO/COM)
    DERIVATIVES_COLUMNS = ['Expiry Date']
    
    # Segment-specific configurations
    SEGMENT_CONFIG = {
        SegmentType.EQUITY: {
            'exchanges': ['NSE', 'BSE'],
            'has_isin': True,
            'has_expiry': False,
            'has_series': True,
            'symbol_pattern': r'^[A-Z0-9]+$',  # Simple alphanumeric
            'product_type': 'CNC'
        },
        SegmentType.FUTURES_OPTIONS: {
            'exchanges': ['NSE', 'BSE'],
            'has_isin': False,
            'has_expiry': True,
            'has_series': False,
            'symbol_pattern': r'^[A-Z]+\d{5}\d+[CP]E$',  # NIFTY2462023300PE
            'product_type': 'NRML'
        },
        SegmentType.COMMODITIES: {
            'exchanges': ['MCX'],
            'has_isin': False,
            'has_expiry': True,
            'has_series': False,
            'symbol_pattern': r'^[A-Z]+\d{2}[A-Z]{3}\d+[CP]E$',  # GOLD24DEC73000PE
            'product_type': 'NRML'
        }
    }
    
    def __init__(self):
        self.data_start_row = 15  # Data starts at row 16 (0-indexed)
        self.header_row = 14      # Header is at row 15 (0-indexed)
    
    def parse_tradebook_file(self, file_path: str) -> pd.DataFrame:
        """
        Parse Zerodha tradebook Excel file with auto-detection of segment type
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            DataFrame with parsed trade data and segment metadata
        """
        try:
            # Validate file exists
            if not Path(file_path).exists():
                raise FileNotFoundError(f"Tradebook file not found: {file_path}")
            
            logger.info(f"Parsing tradebook file: {file_path}")
            
            # Step 1: Auto-detect segment type and file structure
            segment_info = self._detect_segment_type(file_path)
            logger.info(f"Detected segment type: {segment_info['segment_type'].value}")
            
            # Step 2: Read Excel file with appropriate column range
            df = self._read_excel_with_detection(file_path, segment_info)
            
            # Step 3: Validate columns for detected segment
            self._validate_segment_columns(df, segment_info['segment_type'])
            
            # Step 4: Clean and process data
            df = self._clean_data(df, segment_info)
            
            # Step 5: Add metadata
            df['source_file'] = Path(file_path).name
            df['import_timestamp'] = pd.Timestamp.now()
            df['detected_segment'] = segment_info['segment_type'].value
            df['file_exchange'] = segment_info['primary_exchange']
            df['has_expiry_data'] = segment_info['has_expiry']
            
            logger.info(f"Successfully parsed {len(df)} {segment_info['segment_type'].value} trades from {file_path}")
            return df
            
        except Exception as e:
            logger.error(f"Error parsing tradebook file {file_path}: {str(e)}")
            raise
    
    def parse_multiple_files(self, file_paths: List[str]) -> pd.DataFrame:
        """
        Parse multiple tradebook files and combine them
        
        Args:
            file_paths: List of file paths
            
        Returns:
            Combined DataFrame with all trades
        """
        all_trades = []
        
        for file_path in file_paths:
            try:
                trades_df = self.parse_tradebook_file(file_path)
                all_trades.append(trades_df)
            except Exception as e:
                logger.error(f"Failed to parse {file_path}: {str(e)}")
                continue
        
        if not all_trades:
            raise ValueError("No tradebook files could be parsed successfully")
        
        # Combine all dataframes
        combined_df = pd.concat(all_trades, ignore_index=True)
        
        # Sort by execution time
        combined_df = combined_df.sort_values('Order Execution Time')
        
        logger.info(f"Combined {len(combined_df)} trades from {len(all_trades)} files")
        return combined_df
    
    def _detect_segment_type(self, file_path: str) -> Dict:
        """
        Auto-detect the segment type (EQ/FO/COM) by analyzing file content
        
        Returns:
            Dict with segment_type, primary_exchange, has_expiry, sheet_name
        """
        import openpyxl
        
        try:
            # Read workbook to check sheet name and sample data
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            sheet_name = ws.title.lower()
            
            # Check sheet name for obvious indicators
            if 'equity' in sheet_name or 'eq' in sheet_name:
                segment_hint = SegmentType.EQUITY
            elif 'f&o' in sheet_name or 'fo' in sheet_name or 'futures' in sheet_name or 'options' in sheet_name:
                segment_hint = SegmentType.FUTURES_OPTIONS
            elif 'commodity' in sheet_name or 'com' in sheet_name:
                segment_hint = SegmentType.COMMODITIES
            else:
                segment_hint = None
            
            # Sample data from first few rows to analyze
            sample_exchanges = []
            sample_segments = []
            sample_symbols = []
            has_expiry_data = False
            
            # Check for expiry data in column O (15th column)
            for row_num in range(16, min(21, ws.max_row + 1)):
                expiry_cell = ws.cell(row=row_num, column=15).value
                if expiry_cell and isinstance(expiry_cell, str) and len(expiry_cell) >= 8:
                    has_expiry_data = True
                    break
            
            # Sample actual data
            for row_num in range(16, min(26, ws.max_row + 1)):  # Check first 10 data rows
                symbol = ws.cell(row=row_num, column=2).value  # Column B
                exchange = ws.cell(row=row_num, column=5).value  # Column E  
                segment = ws.cell(row=row_num, column=6).value  # Column F
                
                if symbol:
                    sample_symbols.append(str(symbol))
                if exchange:
                    sample_exchanges.append(str(exchange))
                if segment:
                    sample_segments.append(str(segment))
            
            # Determine segment type based on data analysis
            primary_exchange = max(set(sample_exchanges), key=sample_exchanges.count) if sample_exchanges else 'NSE'
            primary_segment = max(set(sample_segments), key=sample_segments.count) if sample_segments else 'EQ'
            
            # Final determination
            if primary_segment == 'EQ':
                detected_segment = SegmentType.EQUITY
            elif primary_segment == 'FO':
                detected_segment = SegmentType.FUTURES_OPTIONS
            elif primary_segment == 'COM' or primary_exchange == 'MCX':
                detected_segment = SegmentType.COMMODITIES
            elif segment_hint:
                detected_segment = segment_hint
            else:
                # Analyze symbol patterns as last resort
                if sample_symbols:
                    # Check for derivatives patterns
                    complex_symbols = [s for s in sample_symbols if len(s) > 10 and any(c.isdigit() for c in s)]
                    if len(complex_symbols) > len(sample_symbols) * 0.5:
                        if primary_exchange == 'MCX':
                            detected_segment = SegmentType.COMMODITIES
                        else:
                            detected_segment = SegmentType.FUTURES_OPTIONS
                    else:
                        detected_segment = SegmentType.EQUITY
                else:
                    detected_segment = SegmentType.EQUITY  # Default
            
            return {
                'segment_type': detected_segment,
                'primary_exchange': primary_exchange,
                'primary_segment': primary_segment,
                'has_expiry': has_expiry_data,
                'sheet_name': ws.title,
                'sample_symbols': sample_symbols[:5],
                'confidence': 'HIGH' if segment_hint == detected_segment else 'MEDIUM'
            }
            
        except Exception as e:
            logger.warning(f"Error in auto-detection, defaulting to EQUITY: {e}")
            return {
                'segment_type': SegmentType.EQUITY,
                'primary_exchange': 'NSE',
                'primary_segment': 'EQ',
                'has_expiry': False,
                'sheet_name': 'Unknown',
                'sample_symbols': [],
                'confidence': 'LOW'
            }
    
    def _read_excel_with_detection(self, file_path: str, segment_info: Dict) -> pd.DataFrame:
        """Read Excel file with appropriate column range based on segment type"""
        
        if segment_info['has_expiry']:
            # FO/COM files have expiry column, need to read wider range
            usecols = 'B:O'  # Include expiry column
        else:
            # EQ files don't have expiry column
            usecols = 'B:N'
        
        df = pd.read_excel(file_path, header=self.header_row, usecols=usecols, engine='openpyxl')
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Handle expiry column naming
        if segment_info['has_expiry'] and len(df.columns) > 13:
            # Rename the last unnamed column to 'Expiry Date'
            last_col = df.columns[-1]
            if 'unnamed' in str(last_col).lower():
                df = df.rename(columns={last_col: 'Expiry Date'})
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        return df
    
    def _validate_segment_columns(self, df: pd.DataFrame, segment_type: SegmentType):
        """Validate columns based on detected segment type"""
        
        missing_columns = []
        config = self.SEGMENT_CONFIG[segment_type]
        
        # Core columns that must be present for all segments
        core_columns = [
            'Symbol', 'Trade Date', 'Exchange', 'Segment', 'Trade Type',
            'Quantity', 'Price', 'Trade ID', 'Order ID', 'Order Execution Time'
        ]
        
        for col in core_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        # Segment-specific validation
        if segment_type == SegmentType.EQUITY:
            # ISIN and Series are important for equity
            if 'ISIN' not in df.columns:
                missing_columns.append('ISIN')
            if 'Series' not in df.columns:
                missing_columns.append('Series')
        
        elif segment_type in [SegmentType.FUTURES_OPTIONS, SegmentType.COMMODITIES]:
            # For derivatives, ISIN and Series are often empty/null, so we don't require them
            # But expiry date is crucial
            if config['has_expiry'] and 'Expiry Date' not in df.columns:
                missing_columns.append('Expiry Date')
        
        if missing_columns:
            raise ValueError(f"Missing required columns for {segment_type.value}: {missing_columns}")
        
        logger.info(f"Column validation passed for {segment_type.value} segment")
    
    def _clean_data(self, df: pd.DataFrame, segment_info: Dict) -> pd.DataFrame:
        """Clean and process tradebook data with segment-specific handling"""
        segment_type = segment_info['segment_type']
        config = self.SEGMENT_CONFIG[segment_type]
        
        # Remove any completely empty rows
        df = df.dropna(how='all')
        
        # Convert data types
        df['Trade Date'] = pd.to_datetime(df['Trade Date'])
        df['Order Execution Time'] = pd.to_datetime(df['Order Execution Time'])
        
        # Convert numeric columns
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce')
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
        df['Trade ID'] = df['Trade ID'].astype(str)
        df['Order ID'] = df['Order ID'].astype(str)
        
        # Handle expiry date for derivatives
        if config['has_expiry'] and 'Expiry Date' in df.columns:
            df['Expiry Date'] = pd.to_datetime(df['Expiry Date'], errors='coerce')
            # Convert to string format for consistency
            df['Expiry Date'] = df['Expiry Date'].dt.strftime('%Y-%m-%d')
            logger.info(f"Processed expiry dates for {segment_type.value} segment")
        
        # Clean string columns
        string_columns = ['Symbol', 'Exchange', 'Segment', 'Trade Type']
        if config['has_series']:
            string_columns.append('Series')
            
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].str.strip().str.upper()
        
        # Handle ISIN based on segment type
        if 'ISIN' in df.columns:
            if config['has_isin']:
                # For equity, ISIN should be present
                df['ISIN'] = df['ISIN'].fillna('').str.strip()
            else:
                # For derivatives, ISIN is typically empty
                df['ISIN'] = df['ISIN'].fillna('').str.strip()
        
        # Segment-specific data cleaning
        if segment_type == SegmentType.EQUITY:
            # For equity, ensure Series is present
            if 'Series' in df.columns:
                df['Series'] = df['Series'].fillna('EQ')
        
        elif segment_type in [SegmentType.FUTURES_OPTIONS, SegmentType.COMMODITIES]:
            # For derivatives, Series is usually empty
            if 'Series' in df.columns:
                df['Series'] = df['Series'].fillna('')
        
        # Standardize trade type
        df['Trade Type'] = df['Trade Type'].str.upper()
        
        # Add derived columns for OrderModel mapping
        df['product_type'] = config['product_type']
        df['segment_type'] = segment_type.value
        
        # Remove rows with invalid data
        df = df[df['Quantity'] > 0]
        df = df[df['Price'] > 0]
        
        # Additional validation for derivatives
        if segment_type in [SegmentType.FUTURES_OPTIONS, SegmentType.COMMODITIES]:
            # Ensure expiry date is present for derivatives
            if 'Expiry Date' in df.columns:
                df = df[df['Expiry Date'].notna()]
        
        logger.info(f"Data cleaning completed for {len(df)} {segment_type.value} trades")
        return df
    
    def get_tradebook_summary(self, df: pd.DataFrame) -> Dict:
        """Get summary statistics of the tradebook"""
        summary = {
            'total_trades': len(df),
            'date_range': {
                'start': df['Trade Date'].min().strftime('%Y-%m-%d'),
                'end': df['Trade Date'].max().strftime('%Y-%m-%d')
            },
            'unique_symbols': df['Symbol'].nunique(),
            'unique_orders': df['Order ID'].nunique(),
            'exchanges': df['Exchange'].unique().tolist(),
            'segments': df['Segment'].unique().tolist(),
            'trade_value': {
                'total': (df['Quantity'] * df['Price']).sum(),
                'buy': (df[df['Trade Type'].str.lower() == 'buy']['Quantity'] * 
                       df[df['Trade Type'].str.lower() == 'buy']['Price']).sum(),
                'sell': (df[df['Trade Type'].str.lower() == 'sell']['Quantity'] * 
                        df[df['Trade Type'].str.lower() == 'sell']['Price']).sum()
            },
            'segment_breakdown': df['Segment'].value_counts().to_dict(),
            'exchange_breakdown': df['Exchange'].value_counts().to_dict()
        }
        
        return summary
    
    def filter_by_date_range(self, df: pd.DataFrame, 
                            start_date: Optional[str] = None,
                            end_date: Optional[str] = None) -> pd.DataFrame:
        """Filter trades by date range"""
        filtered_df = df.copy()
        
        if start_date:
            start_dt = pd.to_datetime(start_date)
            filtered_df = filtered_df[filtered_df['Trade Date'] >= start_dt]
        
        if end_date:
            end_dt = pd.to_datetime(end_date)
            filtered_df = filtered_df[filtered_df['Trade Date'] <= end_dt]
        
        return filtered_df
    
    def group_by_symbol(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Group trades by symbol"""
        return {symbol: group for symbol, group in df.groupby('Symbol')}
    
    def validate_tradebook_integrity(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Validate tradebook data integrity
        
        Returns:
            Tuple of (is_valid, list_of_issues)
        """
        issues = []
        
        # Check for duplicate trade IDs
        duplicate_trades = df[df.duplicated('Trade ID', keep=False)]
        if not duplicate_trades.empty:
            issues.append(f"Found {len(duplicate_trades)} duplicate Trade IDs")
        
        # Check for invalid quantities
        invalid_qty = df[df['Quantity'] <= 0]
        if not invalid_qty.empty:
            issues.append(f"Found {len(invalid_qty)} trades with invalid quantities")
        
        # Check for invalid prices
        invalid_price = df[df['Price'] <= 0]
        if not invalid_price.empty:
            issues.append(f"Found {len(invalid_price)} trades with invalid prices")
        
        # Check for missing symbols
        missing_symbols = df[df['Symbol'].isna() | (df['Symbol'] == '')]
        if not missing_symbols.empty:
            issues.append(f"Found {len(missing_symbols)} trades with missing symbols")
        
        # Check date consistency
        if df['Order Execution Time'].min() < df['Trade Date'].min():
            issues.append("Found execution times before trade dates")
        
        is_valid = len(issues) == 0
        return is_valid, issues
    
    def export_parsed_data(self, df: pd.DataFrame, output_path: str, format: str = 'csv'):
        """Export parsed data to file"""
        if format == 'csv':
            df.to_csv(output_path, index=False)
        elif format == 'excel':
            df.to_excel(output_path, index=False, engine='openpyxl')
        elif format == 'parquet':
            df.to_parquet(output_path, index=False)
        else:
            raise ValueError(f"Unsupported export format: {format}")
        
        logger.info(f"Exported {len(df)} trades to {output_path}")